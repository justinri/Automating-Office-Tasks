#Notes:
# 1. What about multi line pack list and no tracking numbers?
import openpyxl
from pathlib import Path
import emails

def daily_tracking_email(cursor, today, table_Packlist_Detail, table_Packlist_Header, table_SO_Detail):
    ### Sitting up contact varibles
    who_to_contact = "removed..."
    contactEmail = "removed..."
    contactPhoneNumber = "removed..."
    emergency_who_to_contact     = "removed..."
    emergency_contactEmail       = "removed..."
    emergency_contactPhoneNumber = "removed..."
    proto_who_to_contact = "Justin Rittenhouse"
    proto_contactEmail = "removed..."
    proto_contactPhoneNumber = "removed..."
    
    ### Customers to skip
    customerSkip = ["2768A", "2768B", "2768S", "1875", "1877", "5934", "5916U", "5916LT", "5916L",  "5916C",  "5916B",  "5916A",  "5912"]

    ### creating a small data set to look over (today only), i.e., optimizing
    table_Packlist_Header = [item for item in table_Packlist_Header if str(item[6]).split()[0] == str(today) and item[1] not in customerSkip]
    packlistNums = [item[0] for item in table_Packlist_Header]
    table_Packlist_Detail = [item for item in table_Packlist_Detail if item[2] in packlistNums]

    ### Getting a small data set to look over for SO_Details
    SO_DetailNum = [item[6] for item in table_Packlist_Detail]
    table_SO_Detail = [item for item in table_SO_Detail if item[1] in SO_DetailNum]

    ### Getting customer information
    wb_obj = openpyxl.load_workbook("function_files/CustomerList" + ".xlsx") 
    sheet = wb_obj.active
    custInfo = []
    for num, rows in enumerate(sheet.iter_rows()):
        if num == 0: continue
        custNum = sheet.cell(row=num+1, column=2).value
        contactEmail = sheet.cell(row=num+1, column=3).value
        custInfo.append((custNum, contactEmail))

    for row in table_Packlist_Detail:
        packlistNum_Detail = row[2]
        for row2 in table_Packlist_Header:
            packlistNum_header = row2[0]
            if packlistNum_Detail == packlistNum_header:
                for num3, row3 in enumerate(custInfo):
                    if row2[1] != row3[0] and num3+1 == len(custInfo):
                            with open("error_log.txt", "a") as myfile:
                                    myfile.write("{}: Function daily_tracking_email could not find customer: {} \n".format(today, row2[1]))
                    elif row2[1] == row3[0]:
                        for row4 in table_SO_Detail:
                            if row4[1] == row[6]:
                                POnumber = row[8]
                                shipDate = row[6]
                                ourPartNumber = row4[6] 
                                description   = row4[34]
                                quantity = row[15]
                                tracking = row[12]
                                packList = row[2]
                                email_to = row3[1]

                                ### Sending email
                                SMTP_SERVER   = 'smtp.mail.yahoo.com'
                                SMTP_LOGIN    = 'removed...'
                                SMTP_PASSWORD = 'removed...'  # Have to use a app password: https://mail.yahoosmallbusiness.com/accountinfo
                                SENDER_NAME  = 'No_Reply '
                                SENDER_EMAIL = 'removed...'
                                RECIPIENT = email_to
                                SUBJECT   = 'Shipment against PO #: {}'.format(POnumber)
                                CONTENT   = """
                                <p>
                                Good Afternoon, <br> 	<br>
                                
                                We have made a shipment against PO {}, see below for more detail. <br> <br>
                                Ship Date: {} <br>
                                Part Description: {} (Our Part Number: {})  <br> <br>
                                Quantity: {} <br>
                                Tracking #: {} <br>
                                Pack List #: {} <br>
                                <br>
                                <br>
                                Please do not reply to this email; this email box is not monitored. For help or additional information, 
                                please contact {} at {} or by phone at {} for production orders; contact {} at {} or by phone at {} for prototype orders. <br>
                                In case of an emergency contact {} at {} or by phone at {}.
                                </p>
                                <p>
                                Best regards, <br>
                                <br>
                                <b>Glassmaster Controls Company, Inc. </b><br>
                                IATF 16949 Certified <br>
                                http://www.gcontrols.com <br>
                                831 Cobb Ave, Kalamazoo, MI 49007 <br>
                                TEL: (269) 382-2010 <br>
                                FAX: (269) 345-5613 

                                </p>

                                """.format(POnumber, shipDate, description, ourPartNumber, quantity, tracking, packList, 
                                           who_to_contact, contactEmail, contactPhoneNumber, 
                                           proto_who_to_contact, proto_contactEmail, proto_contactPhoneNumber, 
                                           emergency_who_to_contact, emergency_contactEmail, emergency_contactPhoneNumber)
                                #Note: you can use HTML to format the content

                                message = emails.html(
                                    subject=SUBJECT, 
                                    html=CONTENT, 
                                    mail_from=(SENDER_NAME, SENDER_EMAIL)
                                )

                                config = {
                                    'host': SMTP_SERVER,
                                    'timeout': 5,
                                    'ssl': True,
                                    'user': SMTP_LOGIN,
                                    'password': SMTP_PASSWORD
                                }

                            r = message.send(to=RECIPIENT, smtp=config)

                            if not r.success:
                                with open("error_log.txt", "a") as myfile:
                                    myfile.write("{}: Email for packlist: {} did not send. \n".format(today, packList))
                                    
                            continue
    return 0

